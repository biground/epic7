from openpyxl import load_workbook
import multiprocessing as mp

class Hero:
    def __init__(self, dict_data):
        # Virgin data
        self.items = []
        # Load data
        self.formula = dict_data['formula'].split(';')
        self.AT = dict_data['AT']
        self.ATp = dict_data['ATp']
        self.ATa = dict_data['ATa']
        self.DF = dict_data['DF']
        self.DFp = dict_data['DFp']
        self.DFa = dict_data['DFa']
        self.HP = dict_data['HP']
        self.HPp = dict_data['HPp']
        self.HPa = dict_data['HPa']
        self.SP = dict_data['SP']
        self.SPa = dict_data['SPa']
        self.CT = dict_data['CT']
        self.CTa = dict_data['CTa']
        self.CD = dict_data['CD']
        self.CDa = dict_data['CDa']
        self.HT = dict_data['HT']
        self.HTa = dict_data['HTa']
        self.EV = dict_data['EV']
        self.EVa = dict_data['EVa']
        self.TO = dict_data['TO']
        self.TOa = dict_data['TOa']
        self.FB = dict_data['FB']
        self.FBa = dict_data['FBa']
    # To remove all clothes
    def strip(self):
        self.items = []
    # To equip one item
    def equip(self, item):
        self.items.append(item)
    def get_status(self):
        ctr_sets = {
            'AT': 0,
            'DF': 0,
            'HP': 0,
            'SP': 0,
            'CT': 0,
            'CD': 0,
            'HT': 0,
            'EV': 0,
            'TO': 0,
            'FB': 0,
            'IM': 0,
            'BL': 0,
            'FU': 0,
        }
        data = {
            'AT': self.AT,
            'ATp': self.ATp,
            'ATa': self.ATa,
            'DF': self.DF,
            'DFp': self.DFp,
            'DFa': self.DFa,
            'HP': self.HP,
            'HPp': self.HPp,
            'HPa': self.HPa,
            'SP': self.SP,
            'SPa': self.SPa,
            'CT': self.CT,
            'CTa': self.CTa,
            'CD': self.CD,
            'CDa': self.CDa,
            'HT': self.HT,
            'HTa': self.HTa,
            'EV': self.EV,
            'EVa': self.EVa,
            'TO': self.TO,
            'TOa': self.TOa,
            'FB': self.FB,
            'FBa': self.FBa,
        }
        for itm in self.items:
            ctr_sets[itm['set']] += 1
            for attrb in itm['attributes']:
                data[attrb['type']] += attrb['value']
        result = {}
        # AT
        vl = data['AT']
        vl_p = data['ATp']
        vl_a = data['ATa']
        if ctr_sets['AT'] > 3:
            vl_p += 35
        result['AT'] = vl * (100 + vl_p) / 100 + vl_a
        # DF
        vl = data['DF']
        vl_p = data['DFp']
        vl_a = data['DFa']
        if ctr_sets['DF'] > 1:
            vl_p += 15
            if ctr_sets['DF'] > 3:
                vl_p += 15
                if ctr_sets['DF'] > 5:
                    vl_p += 15
        result['DF'] = vl * (100 + vl_p) / 100 + vl_a
        # HP
        vl = data['HP']
        vl_p = data['HPp']
        vl_a = data['HPa']
        if ctr_sets['HP'] > 1:
            vl_p += 15
            if ctr_sets['HP'] > 3:
                vl_p += 15
                if ctr_sets['HP'] > 5:
                    vl_p += 15
        result['HP'] = vl * (100 + vl_p) / 100 + vl_a
        # SP
        vl = data['SP']
        vl_a = data['SPa']
        if ctr_sets['SP'] > 3:
            vl_p = 25
        else:
            vl_p = 0
        result['SP'] = vl * (100 + vl_p) / 100 + vl_a
        # CT
        vl = data['CT'] + data['CTa']
        if ctr_sets['CT'] > 1:
            vl += 12
            if ctr_sets['CT'] > 3:
                vl += 12
                if ctr_sets['CT'] > 5:
                    vl += 12
        if vl > 100:
            vl = 100
        result['CT'] = vl
        # CD
        vl = data['CD'] + data['CDa']
        if ctr_sets['CT'] > 3:
            vl += 40
        result['CD'] = vl
        # HT
        vl = data['HT'] + data['HTa']
        if ctr_sets['HT'] > 1:
            vl += 20
            if ctr_sets['HT'] > 3:
                vl += 20
                if ctr_sets['HT'] > 5:
                    vl += 20
        result['HT'] = vl
        # EV
        vl = data['EV'] + data['EVa']
        if ctr_sets['EV'] > 1:
            vl += 20
            if ctr_sets['EV'] > 3:
                vl += 20
                if ctr_sets['EV'] > 5:
                    vl += 20
        result['EV'] = vl
        # TO
        vl = data['TO'] + data['TOa']
        if ctr_sets['TO'] > 1:
            vl += 4
            if ctr_sets['TO'] > 3:
                vl += 4
                if ctr_sets['TO'] > 5:
                    vl += 4
        result['TO'] = vl
        # FB
        result['FB'] = (ctr_sets['FB'] > 3)
        # FU
        result['FU'] = (ctr_sets['FU'] > 3)
        # IM
        result['IM'] = (ctr_sets['IM'] > 1)
        # BL
        result['BL'] = (ctr_sets['BL'] > 3)
        return result
    # To calc benchmark
    def get_benchmark(self):
        benchmark = []
        status = self.get_status()
        for crtr in self.formula:
            benchmark.append(self.calc_criteria(crtr, status))
        return benchmark
    # To calc on criteria
    def calc_criteria(self, criteria, status):
        if criteria == 'SP':
            return status['SP']
        elif criteria == 'HP':
            return status['HP']
        elif criteria == 'DMG':
            vl = status['AT'] * (100 - status['CT']) + status['AT'] * status['AT'] * status['CT'] * status['CD'] / 100
            if status['FU']:
                return vl * 1.3
            return vl
        elif criteria == 'DPS':
            return self.calc_criteria('DMG', status) * status['SP']
        elif criteria == 'HT':
            return status['HT']
        elif criteria == 'CT':
            return status['CT']
        else:
            print('unexpected criteria {nm}'.format(nm=criteria))
def load_sheet_item(_type, _holder):
    _holder[_type] = []
    ws = WB[_type]
    for rw in ws.rows:
        # 'used' by default False
        wp = {'used': False}
        wp['set'] = rw[0].value
        wp['attributes'] = []
        for i in range(5):
            if rw[i * 2 + 1].value == None:
                continue
            wp['attributes'].append(
                {'type': rw[i * 2 + 1].value, 'value': rw[i * 2 + 2].value}
            )
        _holder[_type].append(wp)
def load_sheet_hero(holder):
    ws = WB['hero']
    idx = 0
    for rw in ws.rows:
        if idx == 0:
            idx += 1
            continue
        hero = Hero({
            'formula': rw[1].value,
            'AT': rw[2].value,
            'ATp': rw[4].value,
            'ATa': rw[3].value,
            'DF': rw[5].value,
            'DFp': 0,
            'DFa': 0,
            'HP': rw[6].value,
            'HPp': rw[7].value,
            'HPa': rw[8].value,
            'SP': rw[9].value,
            'SPa': rw[10].value,
            'CT': rw[11].value,
            'CTa': rw[12].value,
            'CD': rw[13].value,
            'CDa': rw[14].value,
            'HT': rw[15].value,
            'HTa': rw[16].value,
            'EV': rw[17].value,
            'EVa': 0,
            'TO': rw[18].value,
            'TOa': 0,
            'FB': 0,
            'FBa': 0,
        })
        holder.append(hero)
        # Check if equipped
        if rw[19].value is not None:
            pass
        idx += 1
def wear(_piece, _sum):
    # Add set
    _sum['set'][_piece['set']] += 1
    # Add attributes
    for attr in _piece['attributes']:
        _sum[attr['type']] += int(attr['value'])
def calc_benchmark_group(data, hero, idx, total, queue):
    # By default benchmark parameter number is 3
    benchmark_best = [0, 0, 0]
    idx_wp = 0
    for wp in data['weapon']:
        # Fileter by index
        if idx_wp % total != idx:
            idx_wp += 1
            continue
        if wp['used']:
            continue
        idx_hd = 0
        for hd in data['head']:
            if hd['used']:
                continue
            idx_am = 0
            for am in data['armor']:
                if am['used']:
                    continue
                idx_nk = 0
                for nk in data['neck']:
                    if nk['used']:
                        continue
                    idx_rg = 0
                    for rg in data['ring']:
                        if rg['used']:
                            continue
                        idx_sh = 0
                        for sh in data['shoe']:
                            if sh['used']:
                                continue
                            hero.strip()
                            hero.equip(wp)
                            hero.equip(hd)
                            hero.equip(am)
                            hero.equip(nk)
                            hero.equip(rg)
                            hero.equip(sh)
                            benchmark = hero.get_benchmark()
                            # Loop into all parameters
                            for i in range(len(benchmark)):
                                # New is better
                                if benchmark_best[i] < benchmark[i]:
                                    # Use new
                                    benchmark_best = benchmark
                                    set_best = [
                                        idx_wp,
                                        idx_hd,
                                        idx_am,
                                        idx_nk,
                                        idx_rg,
                                        idx_sh,
                                    ]
                                # New is worse
                                elif benchmark_best[i] > benchmark[i]:
                                    # Next please
                                    break
                                # New is same
                                else:
                                    # Go to next parameter
                                    pass
                            idx_sh += 1
                        idx_rg += 1
                    idx_nk += 1
                idx_am += 1
            idx_hd += 1
        idx_wp += 1
    print('core {i} finished with {j} data set'.format(
        i=idx, j=idx_wp * idx_hd * idx_am * idx_nk * idx_rg * idx_sh
    ))
    queue.put({'set_best': set_best, 'benchmark_best': benchmark_best})
if __name__ == '__main__':
    # Open excel data
    # pth_data = r'S:/e7/test.xlsx'
    pth_data = r'D:/SJ/e7/data.xlsx'
    WB = load_workbook(pth_data)
    # Load items data
    items = {}
    # Load weapon data
    load_sheet_item('weapon', items)
    load_sheet_item('head', items)
    load_sheet_item('armor', items)
    load_sheet_item('neck', items)
    load_sheet_item('ring', items)
    load_sheet_item('shoe', items)
    # Load hero data
    heroes = []
    load_sheet_hero(heroes)
    idx_hero = 0
    for hr in heroes:
        # Get formula
        task_number = 10
        processes = []
        # Queue for results
        q = mp.Queue()
        for i in range(task_number):
            process = mp.Process(
                target=calc_benchmark_group,
                args=(items, hr, i, task_number, q)
            )
            processes.append(process)
            process.start()
        for i in range(task_number):
            processes[i].join()
        # Choose best result
        for i in range(task_number):
            benchmark_best = [0, 0, 0]
            result = q.get()
            for j in range(len(result['benchmark_best'])):
                # New is better
                if result['benchmark_best'][j] > benchmark_best[j]:
                    benchmark_best = result['benchmark_best']
                    set_best = result['set_best']
                # New is worse
                elif result['benchmark_best'][j] < benchmark_best[j]:
                    # Give up this result
                    break
                # New is same
                else:
                    # Compare next parameter
                    pass
        # Update excel sheet
        ws = WB['hero']
        ws['T{idx}'.format(idx=idx_hero + 2)] = set_best[0]
        ws['U{idx}'.format(idx=idx_hero + 2)] = set_best[1]
        ws['V{idx}'.format(idx=idx_hero + 2)] = set_best[2]
        ws['W{idx}'.format(idx=idx_hero + 2)] = set_best[3]
        ws['X{idx}'.format(idx=idx_hero + 2)] = set_best[4]
        ws['Y{idx}'.format(idx=idx_hero + 2)] = set_best[5]
        WB.save(pth_data)
        # Set items as 'used'
        items['weapon'][set_best[0]]['used'] = True
        items['head'][set_best[1]]['used'] = True
        items['armor'][set_best[2]]['used'] = True
        items['neck'][set_best[3]]['used'] = True
        items['ring'][set_best[4]]['used'] = True
        items['shoe'][set_best[5]]['used'] = True
        # Equip item on the hero
        hr.strip()
        hr.equip(items['weapon'][set_best[0]])
        hr.equip(items['head'][set_best[1]])
        hr.equip(items['armor'][set_best[2]])
        hr.equip(items['neck'][set_best[3]])
        hr.equip(items['ring'][set_best[4]])
        hr.equip(items['shoe'][set_best[5]])
        # Print result
        print(result['benchmark_best'])
        print(items['weapon'][set_best[0]])
        print(items['head'][set_best[1]])
        print(items['armor'][set_best[2]])
        print(items['neck'][set_best[3]])
        print(items['ring'][set_best[4]])
        print(items['shoe'][set_best[5]])
        idx_hero += 1